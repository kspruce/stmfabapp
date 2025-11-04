# stm_fab/analysis/batch_metrics.py
"""
Batch process metrics analysis for full fabrication sequences
"""
import numpy as np
import pandas as pd
from pathlib import Path
from typing import List, Dict, Any, Optional
from datetime import datetime
import json

from stm_fab.analysis.process_metrics import analyze_process_file, detect_file_type


# Process order for typical fabrication sequence
PROCESS_ORDER = {
    'outgas': 1,
    'degas': 1,
    'flash': 2,
    'termination': 3,
    'hterm': 3,
    'dose': 4,
    'incorporation': 5,
    'inc': 5,
    'overgrowth': 6,
    'susi': 6,
    'unknown': 99
}


def sort_files_by_process_order(files: List[Path]) -> List[Path]:
    """
    Sort files by fabrication process order
    
    Args:
        files: List of file paths
        
    Returns:
        Sorted list of files
    """
    def get_sort_key(filepath: Path):
        file_type = detect_file_type(filepath.name)
        order = PROCESS_ORDER.get(file_type, 99)
        # Secondary sort by filename to handle multiple doses
        return (order, filepath.name)
    
    return sorted(files, key=get_sort_key)


def analyze_folder_metrics(folder_path: str, 
                           calibration: Optional[Dict[str, Any]] = None,
                           locking_layer_rate_ML_min: Optional[float] = None,
                           lte_rate_ML_min: Optional[float] = None,
                           use_active_susi_calibration: bool = True,
                           calibration_db_path: Optional[str] = None,
                           save_json: bool = False,
                           auto_detect_thresholds: bool = False,
                           thresholds: Optional[Dict[str, float]] = None) -> Dict[str, Any]:
    """
    Analyze all LabVIEW files in a folder in process order.

    Fallbacks for growth rates:
    1) Explicit args (highest priority)
    2) Active calibration from DB if use_active_susi_calibration=True
    3) Environment variables SUSI_LOCK_RATE_ML_MIN / SUSI_LTE_RATE_ML_MIN
    """
    import os

    folder = Path(folder_path)
    if not folder.exists():
        return {'error': f'Folder not found: {folder_path}'}

    if (locking_layer_rate_ML_min is None or lte_rate_ML_min is None) and use_active_susi_calibration:
        try:
            from stm_fab.analysis.susi_calibration import SUSICalibrationManager
            manager = SUSICalibrationManager(calibration_db_path)
            active_cal = manager.get_active_calibration()
            if active_cal:
                locking_layer_rate_ML_min = active_cal.get_rate_ML_min('locking_layer')
                lte_rate_ML_min = active_cal.get_rate_ML_min('lte')
                print(f"✓ Loaded active SUSI calibration (ID {active_cal.calibration_id}) from {active_cal.date}")
                print(f"  Method:        {active_cal.method}")
                print(f"  Sample:        {active_cal.sample_name}")
                print(f"  Locking Layer: {locking_layer_rate_ML_min:.3f} ML/min")
                print(f"  LTE:           {lte_rate_ML_min:.3f} ML/min\n")
            else:
                print("⚠ No active SUSI calibration found in database")
        except ImportError:
            print("⚠ susi_calibration module not available")

    # Env var fallback
    if locking_layer_rate_ML_min is None:
        env_lock = os.environ.get('SUSI_LOCK_RATE_ML_MIN')
        if env_lock:
            try:
                locking_layer_rate_ML_min = float(env_lock)
                print(f"Using SUSI_LOCK_RATE_ML_MIN from environment: {locking_layer_rate_ML_min:.3f} ML/min")
            except ValueError:
                pass
    if lte_rate_ML_min is None:
        env_lte = os.environ.get('SUSI_LTE_RATE_ML_MIN')
        if env_lte:
            try:
                lte_rate_ML_min = float(env_lte)
                print(f"Using SUSI_LTE_RATE_ML_MIN from environment: {lte_rate_ML_min:.3f} ML/min")
            except ValueError:
                pass

    # Find all .txt files
    txt_files = list(folder.glob("*.txt"))
    if not txt_files:
        return {'error': f'No .txt files found in {folder_path}'}

    sorted_files = sort_files_by_process_order(txt_files)

    results = []
    process_summary = {
        'folder': str(folder.absolute()),
        'analysis_timestamp': datetime.now().isoformat(),
        'total_files': len(sorted_files),
        'calibration_used': calibration is not None,
        'growth_rates_provided': (locking_layer_rate_ML_min is not None and lte_rate_ML_min is not None),
        'auto_detect_thresholds': bool(auto_detect_thresholds)
    }

    if locking_layer_rate_ML_min is not None:
        process_summary['locking_layer_rate_ML_min'] = locking_layer_rate_ML_min
    if lte_rate_ML_min is not None:
        process_summary['lte_rate_ML_min'] = lte_rate_ML_min

    for i, filepath in enumerate(sorted_files, 1):
        print(f"Analyzing {i}/{len(sorted_files)}: {filepath.name}")
        try:
            result = analyze_process_file(
                str(filepath), 
                calibration=calibration,
                locking_layer_rate_ML_min=locking_layer_rate_ML_min,
                lte_rate_ML_min=lte_rate_ML_min,
                auto_detect_thresholds=auto_detect_thresholds,
                thresholds=thresholds
            )
            result['sequence_number'] = i
            results.append(result)
        except Exception as e:
            results.append({
                'filename': filepath.name,
                'filepath': str(filepath),
                'sequence_number': i,
                'error': str(e)
            })

    process_summary['files'] = results

    stats = generate_process_statistics(results)
    process_summary['statistics'] = stats

    if save_json:
        output_path = folder / f"metrics_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        with open(output_path, 'w') as f:
            json.dump(process_summary, f, indent=2)
        process_summary['saved_to'] = str(output_path)
        print(f"\nSaved results to: {output_path}")

    return process_summary

def generate_process_statistics(results: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Generate summary statistics from analysis results"""
    stats = {
        'total_processes': len(results),
        'successful_analyses': sum(1 for r in results if 'error' not in r),
        'failed_analyses': sum(1 for r in results if 'error' in r)
    }
    
    # Count by type
    type_counts = {}
    for r in results:
        if 'file_type' in r:
            ft = r['file_type']
            type_counts[ft] = type_counts.get(ft, 0) + 1
    stats['process_types'] = type_counts
    
    # Aggregate metrics by type
    stats['totals'] = {}
    
    # Total flash time
    total_flash_time = 0
    flash_count = 0
    for r in results:
        if r.get('file_type') == 'flash' and 'metrics' in r:
            m = r['metrics']
            if 'total_flash_time_s' in m:
                total_flash_time += m['total_flash_time_s']
                flash_count += m.get('flash_count', 0)
    
    if flash_count > 0:
        stats['totals']['total_flash_count'] = flash_count
        stats['totals']['total_flash_time_s'] = total_flash_time
    
    # Total dose exposure
    total_exposure = 0
    dose_count = 0
    for r in results:
        if r.get('file_type') == 'dose' and 'metrics' in r:
            m = r['metrics']
            if m.get('dose_detected') and 'exposure_langmuirs' in m:
                total_exposure += m['exposure_langmuirs']
                dose_count += 1
    
    if dose_count > 0:
        stats['totals']['dose_count'] = dose_count
        stats['totals']['total_exposure_langmuirs'] = total_exposure
    
    # Termination exposure
    for r in results:
        if r.get('file_type') in ['termination', 'hterm'] and 'metrics' in r:
            m = r['metrics']
            if m.get('dose_detected'):
                stats['totals']['termination_exposure_langmuirs'] = m.get('exposure_langmuirs', 0)
                stats['totals']['termination_duration_s'] = m.get('dose_duration_s', 0)
    
    # Outgas metrics
    for r in results:
        if r.get('file_type') == 'outgas' and 'metrics' in r:
            m = r['metrics']
            stats['totals']['outgas_base_pressure_mbar'] = m.get('base_pressure_mbar', 0)
            stats['totals']['outgas_time_to_stabilize_s'] = m.get('time_to_stabilize_s', 0)
    
    # SUSI and overgrowth metrics
    for r in results:
        if r.get('file_type') == 'susi' and 'metrics' in r:
            m = r['metrics']
            if 'susi' in m and 'total_operating_time_s' in m['susi']:
                stats['totals']['susi_operating_time_s'] = m['susi']['total_operating_time_s']
            if 'overgrowth' in m:
                ovg = m['overgrowth']
                if 'total_duration_s' in ovg:
                    stats['totals']['overgrowth_duration_s'] = ovg['total_duration_s']
                if 'total_deposited_ML' in ovg and ovg['total_deposited_ML'] is not None:
                    stats['totals']['total_deposited_ML'] = ovg['total_deposited_ML']
                    stats['totals']['total_deposited_nm'] = ovg.get('total_deposited_nm', 0)
    
    return stats


def format_batch_results(results_dict: Dict[str, Any]) -> str:
    """
    Format batch analysis results for display (text block).
    Shows human-readable durations, per-phase ML and nm, and tags estimated RT.
    """
    def fmt_hms(seconds: float) -> str:
        try:
            seconds = int(round(float(seconds)))
        except Exception:
            return "0 h 0 min 0 s"
        h = seconds // 3600
        m = (seconds % 3600) // 60
        s = seconds % 60
        return f"{h} h {m} min {s} s"

    lines = []
    lines.append("=" * 100)
    lines.append("BATCH PROCESS METRICS ANALYSIS")
    lines.append("=" * 100)
    lines.append("")
    lines.append(f"Folder:       {results_dict['folder']}")
    lines.append(f"Date:         {results_dict['analysis_timestamp']}")
    lines.append(f"Total Files:  {results_dict['total_files']}")
    lines.append(f"Calibration:  {'Yes' if results_dict['calibration_used'] else 'No'}")

    if results_dict.get('growth_rates_provided'):
        lines.append(f"Growth Rates: Yes")
        if 'locking_layer_rate_ML_min' in results_dict:
            lines.append(f"  Locking Layer: {results_dict['locking_layer_rate_ML_min']:.2f} ML/min")
        if 'lte_rate_ML_min' in results_dict:
            lines.append(f"  LTE:           {results_dict['lte_rate_ML_min']:.2f} ML/min")
    else:
        lines.append(f"Growth Rates: No (deposition not calculated)")
    lines.append("")

    # Statistics
    stats = results_dict['statistics']
    lines.append("=" * 100)
    lines.append("SUMMARY STATISTICS")
    lines.append("=" * 100)
    lines.append(f"Successful: {stats['successful_analyses']}/{stats['total_processes']}")
    lines.append(f"Failed:     {stats['failed_analyses']}/{stats['total_processes']}")
    lines.append("")

    if 'process_types' in stats:
        lines.append("Process Types:")
        for ptype, count in sorted(stats['process_types'].items()):
            lines.append(f"  {ptype.upper():15s}: {count}")
        lines.append("")

    if 'totals' in stats and stats['totals']:
        lines.append("Aggregate Metrics:")
        totals = stats['totals']
        if 'total_flash_count' in totals:
            lines.append(f"  Total Flashes:          {totals['total_flash_count']}")
            lines.append(f"  Total Flash Time:       {fmt_hms(totals['total_flash_time_s'])}")
        if 'dose_count' in totals:
            lines.append(f"  Dose Steps:             {totals['dose_count']}")
            lines.append(f"  Total Exposure:         {totals['total_exposure_langmuirs']:.2f} Langmuirs")
        if 'termination_exposure_langmuirs' in totals:
            lines.append(f"  Termination Exposure:   {totals['termination_exposure_langmuirs']:.2f} Langmuirs")
            lines.append(f"  Termination Duration:   {fmt_hms(totals['termination_duration_s'])}")
        if 'outgas_base_pressure_mbar' in totals:
            lines.append(f"  Outgas Base Pressure:   {totals['outgas_base_pressure_mbar']:.2e} mbar")
            if totals.get('outgas_time_to_stabilize_s'):
                lines.append(f"  Outgas Stabilize Time:  {fmt_hms(totals['outgas_time_to_stabilize_s'])}")
        if 'susi_operating_time_s' in totals:
            lines.append(f"  SUSI Operating Time:    {fmt_hms(totals['susi_operating_time_s'])}")
        if 'overgrowth_duration_s' in totals:
            lines.append(f"  Overgrowth Duration:    {fmt_hms(totals['overgrowth_duration_s'])}")
        if 'total_deposited_ML' in totals and totals['total_deposited_ML'] is not None:
            lines.append(f"  Total Silicon Deposited: {totals['total_deposited_ML']:.1f} ML "
                        f"({totals.get('total_deposited_nm', 0):.2f} nm)")
        lines.append("")

    # Individual file results
    lines.append("=" * 100)
    lines.append("PROCESS SEQUENCE (in order)")
    lines.append("=" * 100)
    lines.append("")

    for file_result in results_dict['files']:
        seq = file_result.get('sequence_number', '?')
        filename = file_result.get('filename', 'unknown')
        file_type = file_result.get('file_type', 'unknown')

        lines.append(f"[{seq}] {filename}")
        lines.append(f"    Type: {file_type.upper()}")

        if 'error' in file_result:
            lines.append(f"    Status: ✗ ERROR - {file_result['error']}")
            lines.append("")
            continue

        lines.append(f"    Status: ✓ SUCCESS")

        if 'metrics' in file_result:
            m = file_result['metrics']

            if file_type == 'outgas':
                lines.append(f"    Base Pressure:    {m.get('base_pressure_mbar', 0):.2e} mbar")
                lines.append(f"    Peak Pressure:    {m.get('peak_pressure_mbar', 0):.2e} mbar")
                if m.get('stabilization_detected'):
                    lines.append(f"    Time to Stabilize: {fmt_hms(m.get('time_to_stabilize_s', 0.0))}")
                else:
                    lines.append(f"    Stabilization:    Not detected")

            elif file_type in ['termination', 'hterm']:
                if m.get('dose_detected'):
                    lines.append(f"    Duration:         {fmt_hms(m.get('dose_duration_s', 0.0))}")
                    lines.append(f"    Exposure:         {m.get('exposure_langmuirs', 0):.2f} L")
                    lines.append(f"    Avg Pressure:     {m.get('mean_dose_pressure_torr', 0):.2e} Torr")
                else:
                    lines.append(f"    Status:           No dose detected")

            elif file_type == 'dose':
                if m.get('dose_detected'):
                    mol_weight = m.get('molecular_weight_gmol', 34.0)
                    molecule = 'AsH₃' if mol_weight > 50 else 'PH₃'
                    lines.append(f"    Molecule:         {molecule} ({mol_weight:.2f} g/mol)")
                    lines.append(f"    Duration:         {fmt_hms(m.get('dose_duration_s', 0.0))}")
                    lines.append(f"    Exposure:         {m.get('exposure_langmuirs', 0):.2f} L")
                    lines.append(f"    Integrated Dose:  {m.get('integrated_dose_cm2', 0):.2e} molecules/cm²")
                else:
                    lines.append(f"    Status:           No dose detected")

            elif file_type == 'flash':
                lines.append(f"    Flash Count:      {m.get('flash_count', 0)}")
                lines.append(f"    Total Time:       {fmt_hms(m.get('total_flash_time_s', 0.0))}")
                lines.append(f"    Peak Temp:        {m.get('peak_temperature_C', 0):.1f} °C")

            elif file_type == 'susi':
                # SUSI operating metrics
                if 'susi' in m:
                    susi = m['susi']
                    lines.append(f"    SUSI Time:        {fmt_hms(susi.get('total_operating_time_s', 0.0))}")
                    lines.append(f"    Mean Current:     {susi.get('mean_operating_current_A', 0):.4f} A")

                # Overgrowth metrics
                if 'overgrowth' in m:
                    ovg = m['overgrowth']
                    # Total deposition and duration
                    if ovg.get('total_deposited_ML') is not None:
                        lines.append(f"    Total Deposited:  {ovg['total_deposited_ML']:.1f} ML "
                                     f"({ovg.get('total_deposited_nm', 0):.2f} nm)")
                    # Prefer HMS total duration
                    if 'total_duration_s' in ovg:
                        lines.append(f"    Total Duration:   {fmt_hms(ovg.get('total_duration_s', 0.0))}")
                    else:
                        lines.append(f"    Total Duration:   {ovg.get('total_duration_formatted', '00:00:00')}")

                    lines.append("")
                    # Phases
                    for phase_key, phase_name in [('RT_growth', 'Locking Layer'),
                                                  ('RTA_anneal', 'RTA'),
                                                  ('LTE_growth', 'LTE')]:
                        phase = ovg.get(phase_key, {})
                        if phase.get('detected'):
                            temp_str = ""
                            if 'median_temperature_C' in phase:
                                temp_str = f", {phase['median_temperature_C']:.0f}°C"
                            # Deposition ML and nm and rate
                            depo_str = ""
                            if phase.get('deposited_ML') is not None:
                                ml_val = phase.get('deposited_ML')
                                nm_val = phase.get('deposited_nm')
                                rate_val = phase.get('deposition_rate_ML_min')
                                nm_s = f" ({nm_val:.2f} nm)" if nm_val is not None else ""
                                rate_s = f" @ {rate_val:.2f} ML/min" if rate_val is not None else ""
                                depo_str = f" → {ml_val:.1f} ML{nm_s}{rate_s}"
                            # Duration in HMS
                            dur_hms = fmt_hms(phase.get('duration_s', 0.0))
                            tag = " (estimated 10 ML)" if phase.get('estimated') else ""
                            lines.append(f"    {phase_name:14s}: {dur_hms}, "
                                         f"{phase.get('median_current_A', 0):.4f} A{temp_str}{tag}{depo_str}")
                        else:
                            lines.append(f"    {phase_name:14s}: Not detected")
        lines.append("")

    lines.append("=" * 100)
    if 'saved_to' in results_dict:
        lines.append(f"\nResults saved to: {results_dict['saved_to']}")

    return "\n".join(lines)

def export_batch_results_excel(results_dict: Dict[str, Any], output_path: str):
    """
    Export batch analysis results to Excel
    
    Args:
        results_dict: Results from analyze_folder_metrics
        output_path: Path to save Excel file
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    
    wb = openpyxl.Workbook()
    
    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    # Header
    ws_summary['A1'] = "Batch Process Metrics Analysis"
    ws_summary['A1'].font = Font(size=16, bold=True)
    ws_summary['A2'] = f"Folder: {results_dict['folder']}"
    ws_summary['A3'] = f"Analysis Date: {results_dict['analysis_timestamp']}"
    ws_summary['A4'] = f"Total Files: {results_dict['total_files']}"
    
    # Statistics
    stats = results_dict['statistics']
    row = 6
    ws_summary[f'A{row}'] = "Statistics"
    ws_summary[f'A{row}'].font = Font(bold=True)
    row += 1
    ws_summary[f'A{row}'] = f"Successful Analyses: {stats['successful_analyses']}"
    row += 1
    ws_summary[f'A{row}'] = f"Failed Analyses: {stats['failed_analyses']}"
    row += 2
    
    # Process types
    if 'process_types' in stats:
        ws_summary[f'A{row}'] = "Process Types"
        ws_summary[f'A{row}'].font = Font(bold=True)
        row += 1
        for ptype, count in sorted(stats['process_types'].items()):
            ws_summary[f'A{row}'] = ptype.upper()
            ws_summary[f'B{row}'] = count
            row += 1
        row += 1
    
    # Totals
    if 'totals' in stats and stats['totals']:
        ws_summary[f'A{row}'] = "Aggregate Metrics"
        ws_summary[f'A{row}'].font = Font(bold=True)
        row += 1
        totals = stats['totals']
        for key, value in totals.items():
            ws_summary[f'A{row}'] = key.replace('_', ' ').title()
            ws_summary[f'B{row}'] = value
            row += 1
    
    # Detailed results sheet
    ws_detail = wb.create_sheet("Detailed Results")
    
    headers = ['Seq', 'Filename', 'Type', 'Status', 'Key Metrics']
    for col, header in enumerate(headers, 1):
        cell = ws_detail.cell(1, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    for row, file_result in enumerate(results_dict['files'], 2):
        ws_detail.cell(row, 1, file_result.get('sequence_number', ''))
        ws_detail.cell(row, 2, file_result.get('filename', ''))
        ws_detail.cell(row, 3, file_result.get('file_type', ''))
        
        if 'error' in file_result:
            ws_detail.cell(row, 4, 'ERROR')
            ws_detail.cell(row, 5, file_result['error'])
        else:
            ws_detail.cell(row, 4, 'SUCCESS')
            
            # Add key metrics
            if 'metrics' in file_result:
                m = file_result['metrics']
                file_type = file_result.get('file_type', '')
                
                metrics_str = ""
                if file_type == 'outgas':
                    metrics_str = f"Base: {m.get('base_pressure_mbar', 0):.2e} mbar"
                elif file_type in ['termination', 'hterm'] and m.get('dose_detected'):
                    metrics_str = f"Exposure: {m.get('exposure_langmuirs', 0):.2f} L"
                elif file_type == 'dose' and m.get('dose_detected'):
                    metrics_str = f"Exposure: {m.get('exposure_langmuirs', 0):.2f} L"
                elif file_type == 'flash':
                    metrics_str = f"Flashes: {m.get('flash_count', 0)}, Time: {m.get('total_flash_time_s', 0):.1f} s"
                elif file_type == 'susi' and 'susi' in m:
                    metrics_str = f"SUSI Time: {m['susi'].get('total_operating_time_s', 0):.1f} s"
                
                ws_detail.cell(row, 5, metrics_str)
    
    # Adjust column widths
    for ws in [ws_summary, ws_detail]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output_path)
    print(f"Saved Excel report to: {output_path}")